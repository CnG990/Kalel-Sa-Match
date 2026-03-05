"""
Exports Excel et PDF pour gestionnaires
Rapports de réservations, revenus, statistiques
"""

from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from apps.reservations.models import Reservation
from apps.payments.models import Payment
from django.db.models import Sum, Count, Avg


class ManagerExports:
    """Classe pour générer les exports gestionnaires"""
    
    @staticmethod
    def export_reservations_excel(gestionnaire, date_debut=None, date_fin=None):
        """
        Export Excel des réservations d'un gestionnaire
        """
        # Créer workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Réservations"
        
        # Headers avec style
        headers = [
            'ID', 'Date Réservation', 'Client', 'Téléphone', 'Terrain',
            'Date Début', 'Date Fin', 'Durée (h)', 'Montant (FCFA)',
            'Statut', 'Paiement'
        ]
        
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Récupérer données
        reservations = Reservation.objects.filter(
            terrain__gestionnaire=gestionnaire,
            deleted_at__isnull=True
        )
        
        if date_debut:
            reservations = reservations.filter(created_at__gte=date_debut)
        if date_fin:
            reservations = reservations.filter(created_at__lte=date_fin)
        
        reservations = reservations.select_related('user', 'terrain').order_by('-created_at')
        
        # Remplir données
        row_num = 2
        total_montant = 0
        
        for reservation in reservations:
            try:
                payment = Payment.objects.get(reservation=reservation)
                statut_paiement = payment.get_statut_display()
            except Payment.DoesNotExist:
                statut_paiement = 'Non trouvé'
            
            data = [
                reservation.id,
                reservation.created_at.strftime('%d/%m/%Y %H:%M'),
                reservation.user.nom_complet,
                reservation.telephone or reservation.user.telephone or 'N/A',
                reservation.terrain.nom,
                reservation.date_debut.strftime('%d/%m/%Y %H:%M'),
                reservation.date_fin.strftime('%d/%m/%Y %H:%M'),
                float(reservation.duree_heures),
                float(reservation.montant_total),
                reservation.get_statut_display(),
                statut_paiement
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if reservation.statut in ['confirmee', 'terminee']:
                total_montant += float(reservation.montant_total)
            
            row_num += 1
        
        # Ligne total
        ws.cell(row=row_num, column=8).value = 'TOTAL:'
        ws.cell(row=row_num, column=8).font = Font(bold=True)
        ws.cell(row=row_num, column=9).value = total_montant
        ws.cell(row=row_num, column=9).font = Font(bold=True)
        ws.cell(row=row_num, column=9).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Ajuster largeur colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Sauvegarder en mémoire
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_statistiques_pdf(gestionnaire, mois=None, annee=None):
        """
        Export PDF des statistiques mensuelles
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0066CC'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        if mois and annee:
            titre = f"Rapport Mensuel - {mois}/{annee}"
        else:
            titre = "Rapport Gestionnaire"
        
        elements.append(Paragraph(titre, title_style))
        elements.append(Paragraph(f"Gestionnaire: {gestionnaire.nom_complet}", styles['Normal']))
        elements.append(Paragraph(f"Date: {timezone.now().strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Calculer stats
        now = timezone.now()
        if mois and annee:
            debut_mois = datetime(annee, mois, 1)
            if mois == 12:
                fin_mois = datetime(annee + 1, 1, 1) - timedelta(seconds=1)
            else:
                fin_mois = datetime(annee, mois + 1, 1) - timedelta(seconds=1)
        else:
            debut_mois = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            fin_mois = now
        
        reservations = Reservation.objects.filter(
            terrain__gestionnaire=gestionnaire,
            created_at__gte=debut_mois,
            created_at__lte=fin_mois,
            deleted_at__isnull=True
        )
        
        stats = reservations.aggregate(
            total_reservations=Count('id'),
            revenus_total=Sum('montant_total', filter=reservations.filter(statut__in=['confirmee', 'terminee']).query),
            duree_moyenne=Avg('duree_heures')
        )
        
        # Tableau stats principales
        stats_data = [
            ['Statistique', 'Valeur'],
            ['Réservations totales', stats['total_reservations'] or 0],
            ['Revenus (FCFA)', f"{float(stats['revenus_total'] or 0):,.0f}"],
            ['Durée moyenne (h)', f"{float(stats['duree_moyenne'] or 0):.1f}"],
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(stats_table)
        elements.append(Spacer(1, 20))
        
        # Stats par terrain
        elements.append(Paragraph("Répartition par Terrain", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        terrains = gestionnaire.terrains.filter(deleted_at__isnull=True)
        terrain_data = [['Terrain', 'Réservations', 'Revenus (FCFA)']]
        
        for terrain in terrains:
            resa_terrain = reservations.filter(terrain=terrain)
            revenus = resa_terrain.filter(statut__in=['confirmee', 'terminee']).aggregate(
                total=Sum('montant_total')
            )['total'] or 0
            
            terrain_data.append([
                terrain.nom,
                resa_terrain.count(),
                f"{float(revenus):,.0f}"
            ])
        
        terrain_table = Table(terrain_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        terrain_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgrey, colors.white])
        ]))
        
        elements.append(terrain_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def export_paiements_excel(gestionnaire, date_debut=None, date_fin=None):
        """
        Export Excel des paiements
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Paiements"
        
        headers = [
            'Référence', 'Date', 'Client', 'Terrain', 'Montant (FCFA)',
            'Méthode', 'Statut', 'Transaction ID'
        ]
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        payments = Payment.objects.filter(
            reservation__terrain__gestionnaire=gestionnaire
        ).select_related('reservation', 'reservation__user', 'reservation__terrain')
        
        if date_debut:
            payments = payments.filter(created_at__gte=date_debut)
        if date_fin:
            payments = payments.filter(created_at__lte=date_fin)
        
        payments = payments.order_by('-created_at')
        
        row_num = 2
        total = 0
        
        for payment in payments:
            data = [
                payment.reference,
                payment.created_at.strftime('%d/%m/%Y %H:%M'),
                payment.reservation.user.nom_complet,
                payment.reservation.terrain.nom,
                float(payment.montant),
                payment.get_methode_display(),
                payment.get_statut_display(),
                payment.transaction_id or 'N/A'
            ]
            
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num).value = value
            
            if payment.statut == 'reussi':
                total += float(payment.montant)
            
            row_num += 1
        
        # Total
        ws.cell(row=row_num, column=4).value = 'TOTAL REUSSI:'
        ws.cell(row=row_num, column=4).font = Font(bold=True)
        ws.cell(row=row_num, column=5).value = total
        ws.cell(row=row_num, column=5).font = Font(bold=True)
        ws.cell(row=row_num, column=5).fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


def generate_reservations_excel_response(gestionnaire, date_debut=None, date_fin=None):
    """Helper pour générer HttpResponse Excel"""
    output = ManagerExports.export_reservations_excel(gestionnaire, date_debut, date_fin)
    
    filename = f"reservations_{gestionnaire.id}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def generate_statistiques_pdf_response(gestionnaire, mois=None, annee=None):
    """Helper pour générer HttpResponse PDF"""
    output = ManagerExports.export_statistiques_pdf(gestionnaire, mois, annee)
    
    filename = f"stats_{gestionnaire.id}_{timezone.now().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(output.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
